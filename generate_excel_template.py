"""生成初始数据Excel模板"""
import openpyxl

wb = openpyxl.Workbook()

# ======== Sheet 1: 语料库 ========
ws1 = wb.active
ws1.title = "1-语料库模板"
headers1 = ["术语(中文)", "分类", "释义", "文化注释", "标签", "英语", "日语", "韩语", "西班牙语", "法语", "阿拉伯语"]
ws1.append(headers1)

corpus_data = [
    ["马面裙", "形制", "明代流行的裙装，前后有平幅裙门，两侧打褶",
     "最具辨识度的汉服单品，非'马面'之意，源于裙门形似马面城墙垛口",
     "明制,裙装,标志性", "Mamian Qun (Horse-Face Skirt)", "馬面裙（ばめんぐん）", "마면군", "Falda Mamian", "Jupe Mamian", "تنورة ماميان"],
    ["云肩", "形制", "披在肩部的装饰性衣饰，多为圆形或云朵形",
     "云肩象征吉祥如意，常绣有云纹、花卉，唐代已有雏形",
     "肩饰,装饰", "Cloud Collar (Yunjian)", "雲肩（うんけん）", "운견", "Cuello Nube", "Col Nuage", "ياقة السحابة"],
    ["补服", "形制", "明清官员官服，前胸后背缀有补子表示品级",
     "补子图案严格按品级：文官绣禽、武官绣兽",
     "官服,明清", "Mandarin Square Robe", "補服（ほふく）", "보복", "Toga Mandarina", "Robe Mandarine", "رداء الماندرين"],
    ["云纹", "纹样", "象征高升和如意的传统纹样，形似祥云",
     "云纹起源商周青铜器，汉代广泛用于织物",
     "吉祥纹样,常用", "Cloud Pattern", "雲紋（うんもん）", "운문", "Patron de Nubes", "Motif Nuage", "نمط السحابة"],
    ["缂丝", "工艺", "通经断纬的高级丝织工艺，可织出精细图案",
     "宋代达到顶峰，一寸缂丝一寸金，非遗",
     "非遗,丝织,高级", "Kesi Silk Tapestry", "綴織（けし）", "극사", "Tapiz Kesi", "Tapisserie Kesi", "نسيج كيسي"],
    ["交领", "形制", "汉服基本领型，左右衣襟交叉于胸前",
     "交领右衽是汉服核心特征，区别于少数民族左衽",
     "领型,基础", "Cross-collar (Jiaoling)", "交領（こうりょう）", "교령", "Cuello Cruzado", "Col Croise", "ياقة متقاطعة"],
    ["直裰", "形制", "明代男子日常服装，交领长袍，两侧开衩",
     "直裰为明代士人常服，儒雅简约",
     "男装,明制", "Straight Robe (Zhiduo)", "直綴（じきてつ）", "직철", "Tunica Recta", "Robe Droite", "رداء مستقيم"],
    ["缠枝莲", "纹样", "莲花与枝蔓交织的连续纹样，寓意生生不息",
     "受佛教影响，莲花为圣洁象征",
     "吉祥纹样,佛教", "Lotus Scroll Pattern", "蓮華唐草文", "연화당초문", "Patron Loto Enredadera", "Motif Lotus Entrelace", "نمط اللوتس المتشابك"],
]
for row in corpus_data:
    ws1.append(row)

# 设置列宽
for col in range(1, len(headers1)+1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

# ======== Sheet 2: 合规规则库 ========
ws2 = wb.create_sheet("2-合规规则库模板")
headers2 = ["国家", "区域", "类别", "敏感词(逗号分隔)", "纹样/符号描述", "风险等级", "风险原因", "修改建议"]
ws2.append(headers2)

rules_data = [
    ["沙特阿拉伯", "中东", "宗教禁忌", "猪,猪肉,猪皮,酒精,酒类,十字架,佛像", "",
     "高风险", "含宗教禁忌元素，可能引发穆斯林消费者强烈反感", "移除与伊斯兰教义冲突的元素，替换为几何纹样或植物纹样"],
    ["印度", "南亚", "宗教禁忌", "牛皮,牛肉,牛图案,印度神像", "",
     "高风险", "牛在印度教中为圣物，牛皮制品或牛图案可能造成严重冒犯", "避免使用牛皮材质和相关图案，替换为棉麻或丝绸材质"],
    ["日本", "东亚", "文化冒犯", "龙纹过于张扬,菊花徽章", "菊花为日本皇室纹章",
     "低风险", "注意日本皇室纹章使用规范", "简化龙纹为祥云纹，避免使用菊花徽章作为装饰"],
    ["美国", "北美", "文化挪用", "印第安羽饰,图腾柱,war bonnet", "",
     "高风险", "未经授权使用原住民文化符号属于文化挪用", "与Native American tribes合作获得授权，或使用中国非遗元素替代"],
    ["德国", "欧洲", "政治敏感", "纳粹,万字符,SS标志,希特勒", "卍字纹（左旋）为佛教吉祥纹样",
     "高风险", "万字符在德国/欧洲语境中高度敏感", "如使用佛教卍字纹，必须在注释中明确标注'佛教吉祥纹样，左旋'"],
    ["全球", "全球", "文化冒犯", "辱华,眯眯眼,Chinese virus,支那", "",
     "高风险", "含对华裔群体的歧视性表述", "杜绝任何歧视性内容，使用尊重、平等的表达方式"],
    ["阿联酋", "中东", "宗教禁忌", "裸体,比基尼,短裙,酒,猪", "",
     "高风险", "伊斯兰文化对裸露和酒精有严格限制", "确保服装遮蔽度足够，避免酒类元素"],
    ["法国", "欧洲", "文化冒犯", "贬低法语,贬低法国文化", "",
     "低风险", "法国消费者对自身文化保护意识强", "尊重法国文化，使用正面表达"],
]
for row in rules_data:
    ws2.append(row)

for col in range(1, len(headers2)+1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

# ======== Sheet 3: 区域审美偏好 ========
ws3 = wb.create_sheet("3-区域审美偏好模板")
headers3 = ["区域", "类别", "偏好描述", "权重(0-1)", "备注"]
ws3.append(headers3)

prefs_data = [
    ["北美", "色彩", "红色（需搭配中性色）", 0.7, "纯红色过于浓烈，搭配黑白灰更受欢迎"],
    ["北美", "色彩", "金色（用作点缀）", 0.6, ""],
    ["北美", "风格", "极简东方风", 0.8, "强调线条和轮廓"],
    ["欧洲", "色彩", "深蓝+金色", 0.7, "皇家蓝调受欧洲市场青睐"],
    ["欧洲", "纹样", "花卉纹（简洁版）", 0.7, ""],
    ["欧洲", "风格", "复古宫廷风", 0.6, "可类比维多利亚时期服饰"],
    ["日韩", "色彩", "淡雅粉彩", 0.8, "低饱和度的粉色、淡蓝、浅紫"],
    ["日韩", "风格", "清雅简约", 0.9, "强调侘寂美学"],
    ["东南亚", "色彩", "鲜艳亮色", 0.8, "高饱和度在热带市场更受欢迎"],
    ["东南亚", "纹样", "热带花卉纹", 0.7, ""],
    ["中东", "色彩", "墨绿+金色", 0.7, "低调奢华色系"],
    ["中东", "风格", "端庄典雅", 0.9, "注重遮蔽度与优雅感"],
]
for row in prefs_data:
    ws3.append(row)

for col in range(1, len(headers3)+1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

# ======== Sheet 4: 知识库 ========
ws4 = wb.create_sheet("4-知识库模板")
headers4 = ["类别", "标题(中文)", "内容(中文)", "英语内容", "跨文化类比(北美)", "跨文化类比(日韩)"]
ws4.append(headers4)

kb_data = [
    ["形制", "马面裙的历史与文化",
     "马面裙是明代最具代表性的女裙形制之一，因裙门形似城墙马面而得名。前后各有平幅裙门，两侧打褶，行走时裙褶摆动，极具韵律感。",
     "The Mamian Qun is one of the most iconic women's skirt styles from the Ming Dynasty. It features flat front and back panels with pleated sides.",
     "如维多利亚时期裙撑的东方版本", "与韩服赤古里裙的东方共鸣"],
    ["工艺", "缂丝技艺",
     "缂丝是中国传统丝绸艺术品中的精华，采用通经断纬技法织就，可织出精细的图案和文字。宋代缂丝达到艺术顶峰。",
     "Kesi is the pinnacle of Chinese silk artistry, using discontinuous weft technique to create intricate patterns.",
     "堪比欧洲挂毯工艺", "与西阵织并称东亚丝织双璧"],
]
for row in kb_data:
    ws4.append(row)

for col in range(1, len(headers4)+1):
    ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

# ======== Sheet 5: 使用说明 ========
ws5 = wb.create_sheet("使用说明")
ws5.append(["CCAE引擎初始数据模板 使用说明"])
ws5.append([""])
ws5.append(["1. 按分类填写对应Sheet页"])
ws5.append(["2. 语料库：每行一个术语，填写中文、分类、各语种译文、文化注释"])
ws5.append(["3. 规则库：每行一条规则，敏感词用逗号分隔"])
ws5.append(["4. 审美偏好：每行一条偏好，权重0-1之间"])
ws5.append(["5. 知识库：每行一条目，填写中文和跨文化类比"])
ws5.append([""])
ws5.append(["导入方式："])
ws5.append(["方案A: 通过管理后台API逐条导入"])
ws5.append(["方案B: 使用Python脚本批量导入"])
ws5.append(["方案C: 直接操作SQLite数据库导入"])

ws5.column_dimensions['A'].width = 60

# 保存
output_path = "data/CCAE_初始数据模板.xlsx"
wb.save(output_path)
print(f"[OK] Excel template saved: {output_path}")
